import openpyxl as xl
from openpyxl.chart import BarChart, Reference, PieChart, PieChart3D


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 4).value = "$" + str(sheet.cell(row, 3).value * 0.9)

    values = Reference(sheet,
                       min_row=2,
                       max_row=4,
                       min_col=4,
                       max_col=4)
    chart = PieChart3D()
    chart.add_data(values)
    sheet.add_chart(chart, "A5")

    wb.save("temp.xlsx")


process_workbook("transactions.xlsx")
